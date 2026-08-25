"""Human-readable previews for project resources in the local console."""

from __future__ import annotations

import base64
import csv
import hashlib
import json
import re
import shutil
import subprocess
import tempfile
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from slidecraft.project_resources import resolve_project_resource

TEXT_SUFFIXES = {".txt", ".md", ".markdown", ".csv", ".tsv", ".json", ".yaml", ".yml", ".toml"}
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg", ".bmp"}


def _clean_text(value: str, limit: int = 80_000) -> str:
    value = value.replace("\x00", "").strip()
    return value[:limit] + ("\n\nPreview shortened." if len(value) > limit else "")


def _office_xml_text(path: Path, members: list[str]) -> str:
    parts: list[str] = []
    with zipfile.ZipFile(path) as archive:
        for member in members:
            if member not in archive.namelist():
                continue
            root = ElementTree.fromstring(archive.read(member))
            paragraphs = []
            for paragraph in root.iter():
                if not paragraph.tag.endswith("}p"):
                    continue
                text = "".join(node.text or "" for node in paragraph.iter() if node.tag.endswith("}t"))
                if text.strip():
                    paragraphs.append(text.strip())
            parts.extend(paragraphs)
    return _clean_text("\n\n".join(parts))


def _xlsx_preview(path: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"kind": "unsupported", "message": "Install Slidecraft document support to preview spreadsheets."}
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for sheet in workbook.worksheets[:8]:
        rows = [["" if value is None else str(value) for value in row] for row in sheet.iter_rows(max_row=40, max_col=16, values_only=True)]
        while rows and not any(rows[-1]):
            rows.pop()
        sheets.append({"name": sheet.title, "rows": rows})
    workbook.close()
    return {"kind": "table", "sheets": sheets}


def _office_page_preview(path: Path, page_limit: int = 12) -> dict[str, Any] | None:
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    pdftoppm = shutil.which("pdftoppm")
    if not soffice or not pdftoppm:
        return None
    signature = f"{path}:{path.stat().st_size}:{path.stat().st_mtime_ns}"
    cache = Path(tempfile.gettempdir()) / "slidecraft-previews" / hashlib.sha256(signature.encode()).hexdigest()[:20]
    cache.mkdir(parents=True, exist_ok=True)
    pdf = cache / f"{path.stem}.pdf"
    try:
        if not pdf.exists():
            subprocess.run(
                [soffice, "--headless", "--convert-to", "pdf", "--outdir", str(cache), str(path)],
                check=True,
                capture_output=True,
                timeout=45,
            )
        prefix = cache / "page"
        pages = sorted(cache.glob("page-*.png"), key=lambda item: int(item.stem.split("-")[-1]))
        if not pages:
            subprocess.run(
                [pdftoppm, "-png", "-r", "110", "-f", "1", "-l", str(page_limit), str(pdf), str(prefix)],
                check=True,
                capture_output=True,
                timeout=45,
            )
            pages = sorted(cache.glob("page-*.png"), key=lambda item: int(item.stem.split("-")[-1]))
        images = ["data:image/png;base64," + base64.b64encode(page.read_bytes()).decode("ascii") for page in pages[:page_limit]]
        return {"kind": "pages", "images": images, "page_count": len(images), "truncated": len(pages) > page_limit}
    except (OSError, subprocess.SubprocessError, ValueError):
        return None


def build_path_preview(path: str | Path, metadata: dict[str, Any] | None = None) -> dict[str, Any]:
    metadata = metadata or {}
    path = Path(path).expanduser().resolve()
    suffix = path.suffix.lower()
    base = {
        "name": metadata.get("name", path.name),
        "category": metadata.get("category", "resource"),
        "media_type": metadata.get("media_type"),
        "description": metadata.get("description") or metadata.get("semantic_role") or "",
        "requested_role": metadata.get("requested_role", ""),
        "semantic_role": metadata.get("semantic_role", ""),
        "provenance": metadata.get("provenance", ""),
    }
    if path.name.endswith(".component.json"):
        try:
            manifest = json.loads(path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError):
            manifest = {}
        preview_name = manifest.get("preview", {}).get("image")
        preview_path = (path.parent / preview_name).resolve() if preview_name else None
        if preview_path and not preview_path.is_file():
            preview_path = next((candidate.resolve() for candidate in sorted(path.parent.glob("preview.*")) if candidate.is_file()), None)
        if preview_path and preview_path.is_file() and preview_path.suffix.lower() in IMAGE_SUFFIXES:
            media_type = "image/svg+xml" if preview_path.suffix.lower() == ".svg" else "image/" + preview_path.suffix.lower().lstrip(".").replace("jpg", "jpeg")
            source = f"data:{media_type};base64," + base64.b64encode(preview_path.read_bytes()).decode("ascii")
            return {
                **base,
                "kind": "embedded_image",
                "source": source,
                "component_id": manifest.get("component_id"),
                "implementation_type": manifest.get("implementation", {}).get("type"),
            }
        return {**base, "kind": "structured", "value": manifest}
    if suffix in IMAGE_SUFFIXES:
        return {**base, "kind": "image"}
    if suffix == ".pdf":
        return {**base, "kind": "pdf"}
    if suffix == ".docx":
        page_preview = _office_page_preview(path)
        if page_preview:
            return {**base, **page_preview}
        return {**base, "kind": "document", "text": _office_xml_text(path, ["word/document.xml"])}
    if suffix == ".pptx":
        page_preview = _office_page_preview(path)
        if page_preview:
            return {**base, **page_preview}
        with zipfile.ZipFile(path) as archive:
            slides = sorted(name for name in archive.namelist() if re.fullmatch(r"ppt/slides/slide\d+\.xml", name))
        return {**base, "kind": "document", "text": _office_xml_text(path, slides)}
    if suffix in {".xlsx", ".xlsm"}:
        return {**base, **_xlsx_preview(path)}
    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with path.open(encoding="utf-8-sig", errors="replace", newline="") as stream:
            rows = [row[:16] for _, row in zip(range(40), csv.reader(stream, delimiter=delimiter))]
        return {**base, "kind": "table", "sheets": [{"name": path.stem, "rows": rows}]}
    if suffix in TEXT_SUFFIXES:
        text = path.read_text(encoding="utf-8", errors="replace")
        if suffix == ".json":
            try:
                value = json.loads(text)
                return {**base, "kind": "structured", "value": value}
            except json.JSONDecodeError:
                pass
        return {**base, "kind": "text", "text": _clean_text(text)}
    return {**base, "kind": "unsupported", "message": "This file can be opened from its project folder."}


def build_resource_preview(location: str | Path, resource_id: str) -> dict[str, Any]:
    resource = resolve_project_resource(location, resource_id)
    preview = build_path_preview(resource["resolved_path"], resource)
    return {**preview, "resource_id": resource_id}
