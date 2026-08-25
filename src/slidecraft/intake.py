"""Deck-level multimodal material normalization and source provenance."""

from __future__ import annotations

import csv
import hashlib
import json
import re
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree


def _hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _extract_path(path: Path, modality: str) -> tuple[Any, str]:
    suffix = path.suffix.lower()
    if modality in {"text", "structured_text"} or suffix in {".txt", ".md"}:
        return path.read_text(encoding="utf-8"), "local_text"
    if suffix == ".json":
        return json.loads(path.read_text(encoding="utf-8")), "local_json"
    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as stream:
            return list(csv.reader(stream)), "local_csv"
    if modality == "pdf" or suffix == ".pdf":
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise RuntimeError("Install pypdf to extract PDF materials") from exc
        reader = PdfReader(str(path))
        return [{"page": index + 1, "text": page.extract_text() or ""} for index, page in enumerate(reader.pages)], "local_pypdf"
    if modality == "document" or suffix == ".docx":
        try:
            from docx import Document
        except ImportError as exc:
            raise RuntimeError("Install python-docx to extract Word materials") from exc
        document = Document(str(path))
        return [paragraph.text for paragraph in document.paragraphs if paragraph.text.strip()], "local_python_docx"
    if modality == "spreadsheet" or suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Install openpyxl to extract spreadsheet materials") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        return {
            sheet.title: [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
            for sheet in workbook.worksheets
        }, "local_openpyxl"
    if modality == "presentation" or suffix == ".pptx":
        with zipfile.ZipFile(path) as archive:
            slide_names = sorted(
                (name for name in archive.namelist() if re.fullmatch(r"ppt/slides/slide\d+\.xml", name)),
                key=lambda name: int(re.search(r"(\d+)", Path(name).stem).group(1)),
            )
            slides = []
            for index, name in enumerate(slide_names, start=1):
                root = ElementTree.fromstring(archive.read(name))
                text = [node.text.strip() for node in root.iter() if node.tag.endswith("}t") and node.text and node.text.strip()]
                slides.append({"slide": index, "text": text})
        return slides, "local_office_xml"
    if modality in {"image", "diagram"}:
        from PIL import Image

        with Image.open(path) as image:
            return {"path": str(path), "dimensions_px": list(image.size), "mode": image.mode}, "image_metadata_pending_vlm"
    return {"path": str(path)}, "binary_asset"


def normalize_deck_intake(request: dict[str, Any], base_dir: Path) -> dict[str, Any]:
    atoms: list[dict[str, Any]] = []
    materials: list[dict[str, Any]] = []
    for material in request.get("materials", []):
        path = None
        if material.get("path"):
            candidate = Path(material["path"])
            path = candidate.resolve() if candidate.is_absolute() else (base_dir / candidate).resolve()
            if not path.exists():
                raise FileNotFoundError(path)
        if "content" in material:
            content, adapter = material["content"], "inline"
        elif path:
            content, adapter = _extract_path(path, material["modality"])
        else:
            raise ValueError(f"Material {material['material_id']} has no content or path")
        authority = material.get("authority", "supporting")
        extraction_status = "pending_agent_vision" if adapter == "image_metadata_pending_vlm" else "ready"
        record = {
            **material,
            "path": str(path) if path else None,
            "sha256": _hash(path) if path else None,
            "extraction_adapter": adapter,
            "extraction_status": extraction_status,
        }
        materials.append(record)
        if extraction_status != "ready":
            continue
        values = content if isinstance(content, list) else [content]
        for index, value in enumerate(values, start=1):
            atoms.append({
                "atom_id": f"{material['material_id']}_ATOM_{index:03d}",
                "material_id": material["material_id"],
                "locator": f"item:{index}",
                "modality": material["modality"],
                "value": value,
                "authority": "authoritative" if authority == "authoritative" else "supporting_evidence",
                "required_usage": bool(material.get("required_usage", authority == "authoritative")),
                "provenance": material.get("provenance", "user_upload"),
            })
    clarification_answers = request.get("clarification_answers", {}).get("answers", [])
    for index, answer in enumerate(clarification_answers, start=1):
        if answer.get("resolution") != "user_answered":
            continue
        atoms.append({
            "atom_id": f"CLARIFICATION_ATOM_{index:03d}",
            "material_id": "USER_CLARIFICATIONS",
            "locator": answer["question_id"],
            "modality": "structured_text",
            "value": answer.get("answer"),
            "authority": "authoritative",
            "required_usage": True,
            "provenance": "user_clarification",
            "impact_dimension": answer.get("impact_dimension"),
        })
    constraints = []
    for index, value in enumerate(request.get("constraints", []), start=1):
        if isinstance(value, str):
            text = value
            strength = "hard"
            record = {}
        else:
            text = value["text"]
            strength = value.get("strength", "hard")
            record = value
        constraints.append({
            "constraint_id": record.get("constraint_id", f"DECK_CONSTRAINT_{index:03d}"),
            "text": text,
            "strength": strength,
            "target": record.get("target", "deck"),
            "status": record.get("status", "active"),
            "classification_source": record.get(
                "classification_source",
                "typed_explicit_constraint_field" if isinstance(value, str) else "agent_or_user_authored",
            ),
        })
    pending_material_ids = [item["material_id"] for item in materials if item["extraction_status"] != "ready"]
    return {
        "schema_version": "1.0.0",
        "deck_id": request["deck_id"],
        "source_atoms": atoms,
        "materials": materials,
        "constraint_register": constraints,
        "hard_constraint_ids": [item["constraint_id"] for item in constraints if item["strength"] == "hard"],
        "quality": {
            "material_count": len(materials),
            "source_atom_count": len(atoms),
            "authoritative_atom_count": sum(atom["authority"] == "authoritative" for atom in atoms),
            "pending_extraction_count": len(pending_material_ids),
            "pending_material_ids": pending_material_ids,
            "planning_ready": not pending_material_ids,
        },
    }
